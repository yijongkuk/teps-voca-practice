from __future__ import annotations

import json
import math
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "TEPS_Voca(통합).xlsx"
CONNECTOR_WORKBOOK_PATH = ROOT / "TEPS_Reading_9-10_Connectors.xlsx"
PRONUNCIATION_PATH = ROOT / "pronunciations.json"
MEANING_OVERRIDE_PATH = ROOT / "meaning_overrides.json"
EXAMPLE_OVERRIDE_PATH = ROOT / "example_overrides.json"
EXTERNAL_LIST_PATH = ROOT / "external_word_lists.json"
EXTERNAL_DETAIL_PATH = ROOT / "external_word_details.json"
TOEFL_LIST_PATH = ROOT / "toefl_word_list.json"
OUTPUT_PATH = ROOT / "words-data.js"

FREQUENT_SHEET = "어휘단어장(통합)"
CONNECTOR_SHEET = "접속사 통합"
ROUTINE_CHUNK_COUNT = 10
# The TOEFL book is a 60-day course; four book days make one study chunk, so the
# list rolls over a 15-day cycle instead of the shared 10-day one.
TOEFL_BOOK_DAYS_PER_CHUNK = 4
TOEFL_CHUNK_COUNT = 15
SOURCE_CHUNK_COUNTS = {"toefl": TOEFL_CHUNK_COUNT}
CLOZE_TOKEN_OVERRIDES = {
    "cloak a in the guise of b": "cloak",
    "give ~ a shot": "give",
    "jump the gun": "the",
    "keep one's fingers crossed": "fingers",
    "take ~for a ride": "take",
    "throw one's weight behind": "weight",
}


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_key(value) -> str:
    return clean(value).lower()


def load_pronunciations() -> dict[str, str]:
    if not PRONUNCIATION_PATH.exists():
        return {}

    payload = json.loads(PRONUNCIATION_PATH.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and isinstance(payload.get("entries"), dict):
        return {normalize_key(key): clean(value) for key, value in payload["entries"].items()}
    if isinstance(payload, dict):
        return {normalize_key(key): clean(value) for key, value in payload.items()}
    return {}


def load_meaning_overrides() -> dict[str, str]:
    if not MEANING_OVERRIDE_PATH.exists():
        return {}

    payload = json.loads(MEANING_OVERRIDE_PATH.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and isinstance(payload.get("entries"), dict):
        return {normalize_key(key): clean(value) for key, value in payload["entries"].items()}
    if isinstance(payload, dict):
        return {normalize_key(key): clean(value) for key, value in payload.items()}
    return {}


def load_example_overrides() -> dict[str, dict[str, str]]:
    if not EXAMPLE_OVERRIDE_PATH.exists():
        return {}

    payload = json.loads(EXAMPLE_OVERRIDE_PATH.read_text(encoding="utf-8"))
    raw_entries = payload.get("entries", payload) if isinstance(payload, dict) else {}
    if not isinstance(raw_entries, dict):
        return {}

    entries: dict[str, dict[str, str]] = {}
    for key, value in raw_entries.items():
        if not isinstance(value, dict):
            continue
        entries[normalize_key(key)] = {
            field: clean(value.get(field))
            for field in ("word", "exampleEn", "exampleKo")
            if clean(value.get(field))
        }
    return entries


def load_external_payload(path: Path) -> dict:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else {}


def get_example_override(
    overrides: dict[str, dict[str, str]],
    word_id: str,
    word: str,
) -> dict[str, str]:
    return overrides.get(normalize_key(word_id)) or overrides.get(normalize_key(word)) or {}


def find_cloze_target(term: str, sentence: str) -> tuple[str, int, int] | None:
    if not term or not sentence:
        return None

    override_token = CLOZE_TOKEN_OVERRIDES.get(normalize_key(term))
    if override_token:
        override = re.search(rf"\b{re.escape(override_token)}\b", sentence, flags=re.IGNORECASE)
        if override:
            return override.group(0), override.start(), override.end()

    exact_pattern = re.escape(term).replace(r"\ ", r"\s+")
    exact = re.search(exact_pattern, sentence, flags=re.IGNORECASE)
    if exact:
        return exact.group(0), exact.start(), exact.end()

    if re.fullmatch(r"[A-Za-z]+", term):
        inflections = [rf"{re.escape(term)}(?:s|es|ed|ing)?"]
        if term.lower().endswith("e") and len(term) > 2:
            inflections.append(rf"{re.escape(term[:-1])}(?:ed|ing)")
        if term.lower().endswith("y") and len(term) > 2:
            inflections.append(rf"{re.escape(term[:-1])}(?:ies|ied)")
        inflected = re.search(
            rf"\b(?:{'|'.join(inflections)})\b",
            sentence,
            flags=re.IGNORECASE,
        )
        if inflected:
            return inflected.group(0), inflected.start(), inflected.end()

    tokens = re.findall(r"[A-Za-z][A-Za-z'-]{2,}", term)
    matches = []
    for token in dict.fromkeys(tokens):
        match = re.search(rf"\b{re.escape(token)}\b", sentence, flags=re.IGNORECASE)
        if match:
            matches.append((token, match))
    if matches:
        max_length = max(len(token) for token, _ in matches)
        _, match = max(
            ((token, match) for token, match in matches if len(token) == max_length),
            key=lambda item: item[1].start(),
        )
        return match.group(0), match.start(), match.end()

    return None


def make_cloze(term: str, sentence: str) -> tuple[str, str]:
    target = find_cloze_target(term, sentence)
    if not target:
        return sentence, term

    answer, start, end = target
    return f"{sentence[:start]}____{sentence[end:]}", answer


def assign_routine_chunks(words: list[dict]) -> None:
    chunk_size = max(1, math.ceil(len(words) / ROUTINE_CHUNK_COUNT))
    for index, word in enumerate(words):
        chunk = (index // chunk_size) + 1
        word["chunk"] = min(ROUTINE_CHUNK_COUNT, chunk)


def build_frequent_words() -> list[dict]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook[FREQUENT_SHEET]

    pronunciation_lookup = load_pronunciations()
    meaning_overrides = load_meaning_overrides()
    example_overrides = load_example_overrides()

    words: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rank = row[0] if len(row) > 0 else None
        word = clean(row[2] if len(row) > 2 else None)
        if not rank or not word:
            continue

        word_id = f"F{int(rank):04d}"
        override = get_example_override(example_overrides, word_id, word)
        word = override.get("word", word)

        sheet_meaning = clean(row[3] if len(row) > 3 else None)
        meaning = meaning_overrides.get(normalize_key(word)) or sheet_meaning

        example_cell = row[4] if len(row) > 4 else None
        example_raw = "" if example_cell is None else str(example_cell)
        sheet_example_en, _, sheet_example_ko = example_raw.partition("\n")
        example_en = override.get("exampleEn") or clean(sheet_example_en)
        example_ko = override.get("exampleKo") or clean(sheet_example_ko)
        if example_en:
            cloze, cloze_answer = make_cloze(word, example_en)
        else:
            cloze, cloze_answer = "", ""

        words.append(
            {
                "id": word_id,
                "source": "frequent",
                "sourceLabel": "TEPS 어휘빈출",
                "rank": int(rank),
                "chunk": 0,
                "word": word,
                "meaning": meaning,
                "pronunciation": pronunciation_lookup.get(normalize_key(word), ""),
                "group": clean(row[1] if len(row) > 1 else None),
                "exampleEn": example_en,
                "exampleKo": example_ko,
                "clozeExample": cloze,
                "clozeAnswer": cloze_answer,
                "expression": "",
                "duplicateFileCount": (row[5] if len(row) > 5 else 0) or 0,
                "appearanceCount": (row[6] if len(row) > 6 else 0) or 0,
            }
        )

    assign_routine_chunks(words)
    return words


def build_external_words() -> list[dict]:
    list_payload = load_external_payload(EXTERNAL_LIST_PATH)
    detail_payload = load_external_payload(EXTERNAL_DETAIL_PATH)
    detail_entries = detail_payload.get("entries", {})
    if not isinstance(detail_entries, dict):
        detail_entries = {}

    pronunciation_lookup = load_pronunciations()
    meaning_overrides = load_meaning_overrides()

    def make_card(
        *,
        word_id: str,
        source: str,
        source_label: str,
        rank: int,
        word: str,
        group: str,
        search_terms: list[str] | None = None,
    ) -> dict:
        key = normalize_key(word)
        detail = detail_entries.get(key, {})
        if not isinstance(detail, dict):
            detail = {}
        meaning = meaning_overrides.get(key) or clean(detail.get("meaning"))
        pronunciation = (
            pronunciation_lookup.get(key) or clean(detail.get("pronunciation"))
        )
        example_en = clean(detail.get("exampleEn"))
        example_ko = clean(detail.get("exampleKo"))
        cloze, cloze_answer = (
            make_cloze(word, example_en) if example_en else ("", "")
        )
        return {
            "id": word_id,
            "source": source,
            "sourceLabel": source_label,
            "rank": rank,
            "chunk": 0,
            "word": word,
            "meaning": meaning,
            "pronunciation": pronunciation,
            "group": group,
            "exampleEn": example_en,
            "exampleKo": example_ko,
            "clozeExample": cloze,
            "clozeAnswer": cloze_answer,
            "expression": "",
            "searchTerms": search_terms or [],
            "duplicateFileCount": 0,
            "appearanceCount": 0,
        }

    oxford_words: list[dict] = []
    for rank, entry in enumerate(list_payload.get("oxford5000", []), start=1):
        word = clean(entry.get("word"))
        if not word:
            continue
        levels = "/".join(clean(level) for level in entry.get("levels", []) if clean(level))
        parts = ", ".join(
            clean(part) for part in entry.get("partsOfSpeech", []) if clean(part)
        )
        group = " · ".join(part for part in (levels, parts) if part)
        oxford_words.append(
            make_card(
                word_id=f"O{rank:04d}",
                source="oxford5000",
                source_label="Oxford 5000",
                rank=rank,
                word=word,
                group=group,
            )
        )

    awl_words: list[dict] = []
    for rank, entry in enumerate(list_payload.get("awl", []), start=1):
        word = clean(entry.get("word"))
        if not word:
            continue
        related_forms = [
            clean(form) for form in entry.get("relatedForms", []) if clean(form)
        ]
        awl_words.append(
            make_card(
                word_id=f"A{rank:03d}",
                source="awl",
                source_label="AWL570",
                rank=rank,
                word=word,
                group=f"Sublist {int(entry.get('sublist') or 0)}",
                search_terms=related_forms,
            )
        )

    assign_routine_chunks(oxford_words)
    assign_routine_chunks(awl_words)
    return [*oxford_words, *awl_words]


def build_toefl_words() -> list[dict]:
    payload = load_external_payload(TOEFL_LIST_PATH)
    entries = payload.get("entries", [])
    if not isinstance(entries, list):
        return []

    detail_entries = load_external_payload(EXTERNAL_DETAIL_PATH).get("entries", {})
    if not isinstance(detail_entries, dict):
        detail_entries = {}
    pronunciation_lookup = load_pronunciations()
    meaning_overrides = load_meaning_overrides()
    example_overrides = load_example_overrides()

    words: list[dict] = []
    for rank, entry in enumerate(entries, start=1):
        word = clean(entry.get("word"))
        if not word:
            continue

        word_id = f"T{rank:04d}"
        key = normalize_key(word)
        detail = detail_entries.get(key, {})
        if not isinstance(detail, dict):
            detail = {}

        senses: list[dict] = []
        for sense in entry.get("senses", []):
            if not isinstance(sense, dict):
                continue
            senses.append(
                {
                    "pos": clean(sense.get("pos")),
                    "synonyms": [clean(item) for item in sense.get("synonyms", []) if clean(item)],
                    "exampleEn": clean(sense.get("exampleEn")),
                }
            )
        antonyms = [clean(item) for item in entry.get("antonyms", []) if clean(item)]
        synonyms = []
        for sense in senses:
            for item in sense["synonyms"]:
                if item not in synonyms:
                    synonyms.append(item)

        override = get_example_override(example_overrides, word_id, word)
        example_en = override.get("exampleEn") or next(
            (sense["exampleEn"] for sense in senses if sense["exampleEn"]),
            clean(detail.get("exampleEn")),
        )
        example_ko = override.get("exampleKo") or ""
        cloze, cloze_answer = make_cloze(word, example_en) if example_en else ("", "")

        parts_of_speech = []
        for sense in senses:
            if sense["pos"] and sense["pos"] not in parts_of_speech:
                parts_of_speech.append(sense["pos"])
        book_day = int(entry.get("day") or 1)

        words.append(
            {
                "id": word_id,
                "source": "toefl",
                "sourceLabel": "TOEFLVOCA",
                "rank": rank,
                "chunk": (book_day - 1) // TOEFL_BOOK_DAYS_PER_CHUNK + 1,
                "bookDay": book_day,
                "word": word,
                "meaning": meaning_overrides.get(key) or clean(detail.get("meaning")),
                "pronunciation": (
                    pronunciation_lookup.get(key) or clean(detail.get("pronunciation"))
                ),
                "group": " · ".join(
                    part for part in (f"DAY {book_day:02d}", ", ".join(parts_of_speech)) if part
                ),
                "exampleEn": example_en,
                "exampleKo": example_ko,
                "clozeExample": cloze,
                "clozeAnswer": cloze_answer,
                "expression": "",
                "senses": senses,
                "synonyms": synonyms,
                "antonyms": antonyms,
                "searchTerms": [*synonyms, *antonyms],
                "duplicateFileCount": 0,
                "appearanceCount": 0,
            }
        )

    return words


def build_connector_words() -> list[dict]:
    if not CONNECTOR_WORKBOOK_PATH.exists():
        return []

    workbook = load_workbook(
        CONNECTOR_WORKBOOK_PATH,
        read_only=True,
        data_only=True,
    )
    sheet = workbook[CONNECTOR_SHEET]
    pronunciation_lookup = load_pronunciations()

    words: list[dict] = []
    for rank, row in enumerate(
        sheet.iter_rows(min_row=6, values_only=True),
        start=1,
    ):
        word = clean(row[0] if len(row) > 0 else None)
        if not word:
            continue

        logic = clean(row[1] if len(row) > 1 else None)
        grammar = clean(row[2] if len(row) > 2 else None)
        meaning = clean(row[3] if len(row) > 3 else None)
        usage_note = clean(row[4] if len(row) > 4 else None)
        example_en = clean(row[5] if len(row) > 5 else None)
        example_ko = clean(row[6] if len(row) > 6 else None)
        cloze, cloze_answer = (
            make_cloze(word, example_en) if example_en else ("", "")
        )

        words.append(
            {
                "id": f"C{rank:03d}",
                "source": "connectors",
                "sourceLabel": "TEPS 접속사",
                "rank": rank,
                "chunk": 0,
                "word": word,
                "meaning": meaning,
                "pronunciation": pronunciation_lookup.get(normalize_key(word), ""),
                "group": " · ".join(part for part in (logic, grammar) if part),
                "exampleEn": example_en,
                "exampleKo": example_ko,
                "clozeExample": cloze,
                "clozeAnswer": cloze_answer,
                "expression": "",
                "usageNote": usage_note,
                "searchTerms": [value for value in (logic, grammar) if value],
                "duplicateFileCount": 0,
                "appearanceCount": (row[7] if len(row) > 7 else 0) or 0,
                "correctCount": (row[8] if len(row) > 8 else 0) or 0,
                "sourceLocation": clean(row[9] if len(row) > 9 else None),
            }
        )

    workbook.close()
    assign_routine_chunks(words)
    return words


def build_words() -> list[dict]:
    return [
        *build_frequent_words(),
        *build_connector_words(),
        *build_external_words(),
        *build_toefl_words(),
    ]


def main() -> None:
    words = build_words()
    sources = ("frequent", "connectors", "oxford5000", "awl", "toefl")
    chunk_counts = {
        source: {
            str(chunk): sum(
                1
                for word in words
                if word["source"] == source and word["chunk"] == chunk
            )
            for chunk in range(
                1, SOURCE_CHUNK_COUNTS.get(source, ROUTINE_CHUNK_COUNT) + 1
            )
        }
        for source in sources
    }
    meaning_coverage = sum(1 for word in words if word["meaning"])
    pronunciation_coverage = sum(1 for word in words if word["pronunciation"])

    meta = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": "src/TEPS_Voca(통합).xlsx",
        "sourceFiles": [
            "src/TEPS_Voca(통합).xlsx",
            "src/TEPS_Reading_9-10_Connectors.xlsx",
            "src/external_word_lists.json",
            "src/external_word_details.json",
            "src/toefl_word_list.json",
        ],
        "total": len(words),
        "counts": {
            source: sum(1 for word in words if word["source"] == source)
            for source in sources
        },
        "coverage": {
            "meaning": meaning_coverage,
            "pronunciation": pronunciation_coverage,
        },
        "chunkCounts": chunk_counts,
        "chunkCountsBySource": {
            source: SOURCE_CHUNK_COUNTS.get(source, ROUTINE_CHUNK_COUNT)
            for source in sources
        },
        "chunkRule": {
            "routine": (
                "split each source list independently and evenly across a "
                f"{ROUTINE_CHUNK_COUNT}-day cycle"
            ),
            "toefl": (
                f"group {TOEFL_BOOK_DAYS_PER_CHUNK} book days into one chunk for a "
                f"{TOEFL_CHUNK_COUNT}-day cycle"
            ),
        },
    }

    js = (
        "// Generated from the TEPS workbook and external word-list data "
        "by src/generate_words_data.py\n"
        f"window.TEPS_META = {json.dumps(meta, ensure_ascii=False, indent=2)};\n"
        f"window.TEPS_WORDS = {json.dumps(words, ensure_ascii=False, indent=2)};\n"
    )
    OUTPUT_PATH.write_text(js, encoding="utf-8")
    print(f"Wrote {len(words)} words to {OUTPUT_PATH}")
    print(f"Meaning coverage: {meaning_coverage}/{len(words)}")
    print(f"Pronunciation coverage: {pronunciation_coverage}/{len(words)}")


if __name__ == "__main__":
    main()
